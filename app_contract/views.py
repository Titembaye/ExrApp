from _decimal import Decimal
from django.core.paginator import Paginator
from django.db.models.functions import ExtractMonth
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from app_client.models import Client
from app_dommage.models import Other
from app_person.models import Person
from app_setting.models import Company, Type, Guaranttee
from app_vehicle.models import Vehicle
from .forms import ContractForm
from .models import Contract
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from django.db.models import Count, Sum, Q, IntegerField
from openpyxl import Workbook
from datetime import timedelta, date
import datetime
from django.http import JsonResponse
from django.db.models import F

from .templatetags.custom_filters import calculate_percentage
from django.db.models import Count, Sum, ExpressionWrapper, DecimalField, Case, When, Value


def contract_list(request):
    contracts_list = Contract.objects.all()

    # Filtres
    policy = request.GET.get('policy')

    if policy:
        contracts_list = contracts_list.filter(policy_num__icontains=policy)

    # Tri par date d'enregistrement décroissante
    contracts_list = contracts_list.order_by('-register_date')
    contracts_count = Contract.objects.all().count()
    paginator = Paginator(contracts_list, 5)
    page = request.GET.get('page')
    contracts = paginator.get_page(page)

    return render(request, 'insurances/contract_list.html',
                  {'contracts': contracts, 'navbar': 'contract', 'contracts_count': contracts_count})


def contract_add(request, vehicle_id=None, other_id=None, person_id=None):
    if request.method == 'POST':
        form = ContractForm(request.POST)

        if form.is_valid():
            new_contract = form.save(commit=False)

            # Effectuez les calculs nécessaires pour les champs dynamiques ici
            new_contract.prime_ht = new_contract.prime_ttc / Decimal('1.06')
            new_contract.committee = (new_contract.prime_net * new_contract.rate) / Decimal('100')
            new_contract.detention = new_contract.committee * Decimal('0.05')
            new_contract.final_com = new_contract.committee - new_contract.detention

            if person_id is not None:
                try:
                    person = Person.objects.get(pk=person_id)
                    new_contract.person = person  # Préremplir le champ person avec l'ID de la personne
                except Person.DoesNotExist:
                    pass
            else:
                if vehicle_id is not None:
                    try:
                        vehicle = Vehicle.objects.get(pk=vehicle_id)
                        new_contract.vehicle = vehicle  # Préremplir le champ vehicle avec l'ID du véhicule
                    except Vehicle.DoesNotExist:
                        pass

            new_contract.save()
            return redirect('app_contract:contract_list')  # Redirigez vers la liste de contrats

    else:
        initial_data = {}  # Initialiser un dictionnaire vide pour les données initiales

        if person_id is not None:
            initial_data['person'] = person_id  # Préremplir le champ person avec l'ID de la personne
        elif vehicle_id is not None:
            initial_data['vehicle'] = vehicle_id  # Préremplir le champ vehicle avec l'ID du véhicule
        elif other_id is not None:
            initial_data['other'] = other_id

        form = ContractForm(initial=initial_data)

    return render(request, 'insurances/contract_form.html', {'form': form})


# -----------------------------------------------------------

def contract_edit(request, pk):
    contract = get_object_or_404(Contract, pk=pk)

    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            form.save()
            return redirect('app_contract:contract_list')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'insurances/contract_form.html',
                  {'form': form,
                   'contract': contract}
                  )


def contract_detail(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    vehicle = contract.vehicle
    person = contract.person
    other = contract.other
    return render(request, 'insurances/contract_detail.html',
                  {'contract': contract, 'vehicle': vehicle, 'person': person, 'other': other})


def export_data(request):
    # Créez un classeur Excel
    wb = Workbook()

    # Feuille 1 : Garantie
    ws1 = wb.create_sheet(title="Garantie")
    ws1.append(['Id', 'Libelle'])
    guarantties = Guaranttee.objects.all()
    for guaranttee in guarantties:
        ws1.append([guaranttee.id, guaranttee.libelle])

    # Feuille 2 : Client
    ws2 = wb.create_sheet(title="Client")
    ws2.append(['Id', 'Genre', 'Nom et prénoms', 'Téléphone 1', 'Téléphone 2'])
    clients = Client.objects.all()
    for client in clients:
        ws2.append([client.id, client.gender, client.full_name, client.phone1, client.phone2])

    # Feuille 3 : Véhicule
    ws3 = wb.create_sheet(title="Véhicule")
    ws3.append(['Id', 'Marque', 'Immatriculation', 'Puissance', 'Sièges', 'Année'])
    vehicles = Vehicle.objects.all()
    for vehicle in vehicles:
        ws3.append([
            vehicle.id, vehicle.brand, vehicle.registration, vehicle.power, vehicle.seats, vehicle.model_year])

    # Feuille 4 : Person
    ws4 = wb.create_sheet(title="Person")
    ws4.append(['Id', 'sexe', 'Nom et Prénoms', 'Naissance'])
    persons = Person.objects.all()
    for person in persons:
        ws4.append([person.id,  person.gender, person.full_name, person.birthday])

    # Feuille 5 : Other
    ws5 = wb.create_sheet(title="Dommage")
    ws5.append(['Id', 'Description', 'Commentaire'])
    others = Other.objects.all()
    for other in others:
        ws5.append([other.id, other.description, other.comment])

    ws6 = wb.create_sheet(title="Sociétés")
    ws6.append(['Id', 'Nom'])
    companies = Company.objects.all()
    for company in companies:
        ws6.append([company.id, company.name])

    ws7 = wb.create_sheet(title="Types")
    ws7.append(['Id', 'Libelle'])
    types_contracts = Type.objects.all()
    for type_c in types_contracts:
        ws7.append([type_c.id, type_c.libelle])

    # Feuille 6 : Contract
    ws8 = wb.create_sheet(title="Contract")
    ws8.append(['Id', 'Numéro de police', 'Date enregistrement', 'Date de debut', 'Echéance', 'Avenant', 'Type contrat',
                'Client', 'Garantie', 'Compagnie', 'Véhicule', 'Personne', 'Dommage', 'Agent', 'Prime ttc',
                'Prime ht', 'Prime nette', 'Commission', 'Retenue', 'Commission finale', 'Taux', 'Accessoires',
                'Commentaire'])
    contracts = Contract.objects.all()
    for contract in contracts:
        client_id = contract.client.id if contract.client else None
        guaranttee_id = contract.guaranttee.id if contract.guaranttee else None
        company_id = contract.company.id if contract.company else None
        type_contract_id = contract.type_contract.id if contract.type_contract else None
        vehicle_id = contract.vehicle.id if contract.vehicle else None
        other_id = contract.other.id if contract.other else None
        person_id = contract.person.id if contract.person else None

        ws8.append([contract.id, contract.policy_num, contract.register_date, contract.effect, contract.due_date
                       , contract.amendment, type_contract_id, client_id, guaranttee_id, company_id, vehicle_id,
                    person_id,
                    other_id, contract.agent, contract.prime_ttc, contract.prime_ht, contract.prime_net,
                    contract.committee, contract.detention, contract.final_com, contract.rate, contract.accessory,
                    contract.assured
                    ])

        # Enregistrez le fichier Excel
    response = HttpResponse(content_type='application/ms-excel')
    response[

        'Content-Disposition'] = 'attachment; filename="export_data.xlsx"'
    wb.save(response)

    return response


def export_data_sanlam(request):
    sanlam = get_object_or_404(Company, name='SANLAM')
    # Créez un classeur Excel
    wb = Workbook()

    # Feuille 1 : Garantie
    ws1 = wb.create_sheet(title="Garantie")
    ws1.append(['Id', 'Libelle'])
    guarantties = Guaranttee.objects.filter(contract__company=sanlam).distinct()
    for guaranttee in guarantties:
        ws1.append([guaranttee.id, guaranttee.libelle])

    # Feuille 2 : Client
    ws2 = wb.create_sheet(title="Client")
    ws2.append(['Id', 'Genre', 'Nom et prénoms', 'Téléphone 1', 'Téléphone 2'])
    clients = Client.objects.filter(contract__company=sanlam)
    for client in clients:
        ws2.append([client.id, client.gender, client.full_name, client.phone1, client.phone2])

    # Feuille 3 : Véhicule
    ws3 = wb.create_sheet(title="Véhicule")
    ws3.append(['Id', 'Marque', 'Immatriculation', 'Puissance', 'Sièges', 'Année'])
    vehicles = Vehicle.objects.filter(contract__company=sanlam)
    for vehicle in vehicles:
        ws3.append([
            vehicle.id, vehicle.brand, vehicle.registration, vehicle.power, vehicle.seats, vehicle.model_year])

    # Feuille 4 : Person
    ws4 = wb.create_sheet(title="Person")
    ws4.append(['Id', 'sexe', 'Nom et Prénoms', 'Naissance'])
    persons = Person.objects.filter(contract__company=sanlam)
    for person in persons:
        ws4.append([person.id,  person.gender, person.full_name, person.birthday])

    # Feuille 5 : Other
    ws5 = wb.create_sheet(title="Dommage")
    ws5.append(['Id', 'Description', 'Commentaire'])
    others = Other.objects.filter(contract__company=sanlam)
    for other in others:
        ws5.append([other.id, other.description, other.comment])

    ws6 = wb.create_sheet(title="Types")
    ws6.append(['Id', 'Libelle'])
    types_contracts = Type.objects.filter(contract__company=sanlam)
    for type_c in types_contracts:
        ws6.append([type_c.id, type_c.libelle])

    # Feuille 6 : Contract
    ws7 = wb.create_sheet(title="Contract")
    ws7.append(['Id', 'Numéro de police', 'Date enregistrement', 'Date de debut', 'Echéance', 'Avenant', 'Type contrat',
                'Client', 'Garantie', 'Compagnie', 'Véhicule', 'Personne', 'Dommage', 'Agent', 'Prime ttc',
                'Prime ht', 'Prime nette', 'Commission', 'Retenue', 'Commission finale', 'Taux', 'Accessoires',
                'Commentaire'])
    contracts = Contract.objects.filter(company=sanlam)
    for contract in contracts:
        client_id = contract.client.id if contract.client else None
        guaranttee_id = contract.guaranttee.id if contract.guaranttee else None
        company_id = contract.company.id if contract.company else None
        type_contract_id = contract.type_contract.id if contract.type_contract else None
        vehicle_id = contract.vehicle.id if contract.vehicle else None
        other_id = contract.other.id if contract.other else None
        person_id = contract.person.id if contract.person else None

        ws7.append([contract.id, contract.policy_num, contract.register_date, contract.effect, contract.due_date
                       , contract.amendment, type_contract_id, client_id, guaranttee_id, company_id, vehicle_id,
                    person_id,
                    other_id, contract.agent, contract.prime_ttc, contract.prime_ht, contract.prime_net,
                    contract.committee, contract.detention, contract.final_com, contract.rate, contract.accessory,
                    contract.assured
                    ])

        # Enregistrez le fichier Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="sanlam_contracts.xlsx"'
    wb.save(response)

    return response

def export_data_nsia(request):
    nsia = get_object_or_404(Company, name='NSIA')
    # Créez un classeur Excel
    wb = Workbook()

    # Feuille 1 : Garantie
    ws1 = wb.create_sheet(title="Garantie")
    ws1.append(['Id', 'Libelle'])
    guarantties = Guaranttee.objects.filter(contract__company=nsia).distinct()
    for guaranttee in guarantties:
        ws1.append([guaranttee.id, guaranttee.libelle])

    # Feuille 2 : Client
    ws2 = wb.create_sheet(title="Client")
    ws2.append(['Id', 'Genre', 'Nom et prénoms', 'Téléphone 1', 'Téléphone 2'])
    clients = Client.objects.filter(contract__company=nsia)
    for client in clients:
        ws2.append([client.id, client.gender, client.full_name, client.phone1, client.phone2])

    # Feuille 3 : Véhicule
    ws3 = wb.create_sheet(title="Véhicule")
    ws3.append(['Id', 'Marque', 'Immatriculation', 'Puissance', 'Sièges', 'Année'])
    vehicles = Vehicle.objects.filter(contract__company=nsia)
    for vehicle in vehicles:
        ws3.append([
            vehicle.id, vehicle.brand, vehicle.registration, vehicle.power, vehicle.seats, vehicle.model_year])

    # Feuille 4 : Person
    ws4 = wb.create_sheet(title="Person")
    ws4.append(['Id', 'sexe', 'Nom et Prénoms', 'Naissance'])
    persons = Person.objects.filter(contract__company=nsia)
    for person in persons:
        ws4.append([person.id,  person.gender, person.full_name, person.birthday])

    # Feuille 5 : Other
    ws5 = wb.create_sheet(title="Dommage")
    ws5.append(['Id', 'Description', 'Commentaire'])
    others = Other.objects.filter(contract__company=nsia)
    for other in others:
        ws5.append([other.id, other.description, other.comment])

    ws6 = wb.create_sheet(title="Types")
    ws6.append(['Id', 'Libelle'])
    types_contracts = Type.objects.filter(contract__company=nsia)
    for type_c in types_contracts:
        ws6.append([type_c.id, type_c.libelle])

    # Feuille 6 : Contract
    ws7 = wb.create_sheet(title="Contract")
    ws7.append(['Id', 'Numéro de police', 'Date enregistrement', 'Date de debut', 'Echéance', 'Avenant', 'Type contrat',
                'Client', 'Garantie', 'Compagnie', 'Véhicule', 'Personne', 'Dommage', 'Agent', 'Prime ttc',
                'Prime ht', 'Prime nette', 'Commission', 'Retenue', 'Commission finale', 'Taux', 'Accessoires',
                'Commentaire'])
    contracts = Contract.objects.filter(company=nsia)
    for contract in contracts:
        client_id = contract.client.id if contract.client else None
        guaranttee_id = contract.guaranttee.id if contract.guaranttee else None
        company_id = contract.company.id if contract.company else None
        type_contract_id = contract.type_contract.id if contract.type_contract else None
        vehicle_id = contract.vehicle.id if contract.vehicle else None
        other_id = contract.other.id if contract.other else None
        person_id = contract.person.id if contract.person else None

        ws7.append([contract.id, contract.policy_num, contract.register_date, contract.effect, contract.due_date
                       , contract.amendment, type_contract_id, client_id, guaranttee_id, company_id, vehicle_id,
                    person_id,
                    other_id, contract.agent, contract.prime_ttc, contract.prime_ht, contract.prime_net,
                    contract.committee, contract.detention, contract.final_com, contract.rate, contract.accessory,
                    contract.assured
                    ])

        # Enregistrez le fichier Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="nsia_contracts.xlsx"'
    wb.save(response)

    return response


def export_data_fidelia(request):
    fidelia = get_object_or_404(Company, name='FIDELIA')
    # Créez un classeur Excel
    wb = Workbook()

    # Feuille 1 : Garantie
    ws1 = wb.create_sheet(title="Garantie")
    ws1.append(['Id', 'Libelle'])
    guarantties = Guaranttee.objects.filter(contract__company=fidelia).distinct()
    for guaranttee in guarantties:
        ws1.append([guaranttee.id, guaranttee.libelle])

    # Feuille 2 : Client
    ws2 = wb.create_sheet(title="Client")
    ws2.append(['Id', 'Genre', 'Nom et prénoms', 'Téléphone 1', 'Téléphone 2'])
    clients = Client.objects.filter(contract__company=fidelia)
    for client in clients:
        ws2.append([client.id, client.gender, client.full_name, client.phone1, client.phone2])

    # Feuille 3 : Véhicule
    ws3 = wb.create_sheet(title="Véhicule")
    ws3.append(['Id', 'Marque', 'Immatriculation', 'Puissance', 'Sièges', 'Année'])
    vehicles = Vehicle.objects.filter(contract__company=fidelia)
    for vehicle in vehicles:
        ws3.append([
            vehicle.id, vehicle.brand, vehicle.registration, vehicle.power, vehicle.seats, vehicle.model_year])

    # Feuille 4 : Person
    ws4 = wb.create_sheet(title="Person")
    ws4.append(['Id', 'sexe', 'Nom et Prénoms', 'Naissance'])
    persons = Person.objects.filter(contract__company=fidelia)
    for person in persons:
        ws4.append([person.id,  person.gender, person.full_name, person.birthday])

    # Feuille 5 : Other
    ws5 = wb.create_sheet(title="Dommage")
    ws5.append(['Id', 'Description', 'Commentaire'])
    others = Other.objects.filter(contract__company=fidelia)
    for other in others:
        ws5.append([other.id, other.description, other.comment])

    ws6 = wb.create_sheet(title="Types")
    ws6.append(['Id', 'Libelle'])
    types_contracts = Type.objects.filter(contract__company=fidelia)
    for type_c in types_contracts:
        ws6.append([type_c.id, type_c.libelle])

    # Feuille 6 : Contract
    ws7 = wb.create_sheet(title="Contract")
    ws7.append(['Id', 'Numéro de police', 'Date enregistrement', 'Date de debut', 'Echéance', 'Avenant', 'Type contrat',
                'Client', 'Garantie', 'Compagnie', 'Véhicule', 'Personne', 'Dommage', 'Agent', 'Prime ttc',
                'Prime ht', 'Prime nette', 'Commission', 'Retenue', 'Commission finale', 'Taux', 'Accessoires',
                'Commentaire'])
    contracts = Contract.objects.filter(company=fidelia)
    for contract in contracts:
        client_id = contract.client.id if contract.client else None
        guaranttee_id = contract.guaranttee.id if contract.guaranttee else None
        company_id = contract.company.id if contract.company else None
        type_contract_id = contract.type_contract.id if contract.type_contract else None
        vehicle_id = contract.vehicle.id if contract.vehicle else None
        other_id = contract.other.id if contract.other else None
        person_id = contract.person.id if contract.person else None

        ws7.append([contract.id, contract.policy_num, contract.register_date, contract.effect, contract.due_date
                       , contract.amendment, type_contract_id, client_id, guaranttee_id, company_id, vehicle_id,
                    person_id,
                    other_id, contract.agent, contract.prime_ttc, contract.prime_ht, contract.prime_net,
                    contract.committee, contract.detention, contract.final_com, contract.rate, contract.accessory,
                    contract.assured
                    ])

        # Enregistrez le fichier Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="fidelia_contracts.xlsx"'
    wb.save(response)

    return response


def expiring_contracts_list(request):
    # Calcule la date actuelle plus un mois

    expiring_contracts = Contract.objects.filter(due_date__lte=datetime.datetime.now() + timedelta(days=30)).order_by(
        '-register_date')
    expiring_contracts_count = Contract.objects.filter(
        due_date__lte=datetime.datetime.now() + timedelta(days=30)).count()
    paginator = Paginator(expiring_contracts, 5)
    page = request.GET.get('page')
    expiring_contracts = paginator.get_page(page)

    return render(request, 'insurances/expiring_contracts.html',
                  {'expiring_contracts': expiring_contracts,
                   'expiring_contracts_count': expiring_contracts_count})


def expired_contracts_list(request):
    # Récupérer la date actuelle
    current_date = date.today()

    # Récupérer la liste des contrats expirés
    expired_contracts = Contract.objects.filter(due_date__lt=current_date).order_by('-register_date')
    expired_contracts_count = Contract.objects.filter(due_date__lt=current_date).count()
    paginator = Paginator(expired_contracts, 5)
    page = request.GET.get('page')
    expired_contracts = paginator.get_page(page)
    return render(request, 'insurances/expired_contract.html',
                  {'expired_contracts': expired_contracts,
                   'expired_contracts_count': expired_contracts_count})


def contract_delete(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('app_contract:contract_list')
    return render(request, 'insurances/contract_confirm_delete.html', {'contract': contract})


def contracts_dashboard_prime(request):
    contracts = Contract.objects.all()

    # Créez des dictionnaires vides pour stocker les primes par mois
    prime_ht_by_month = {}
    prime_nette_by_month = {}

    # Parcourez les contrats et ajoutez les primes au dictionnaire correspondant
    for contract in contracts:
        register_date = contract.register_date
        if register_date:
            register_month = register_date.month
            prime_ht = contract.prime_ht
            prime_nette = contract.prime_net

            # Si le mois n'existe pas encore dans le dictionnaire, créez-le
            if register_month not in prime_ht_by_month:
                prime_ht_by_month[register_month] = 0
                prime_nette_by_month[register_month] = 0

            # Ajoutez les primes au mois correspondant
            prime_ht_by_month[register_month] += prime_ht
            prime_nette_by_month[register_month] += prime_nette

    # Convertissez les dictionnaires en listes pour les utiliser dans Chart.js
    months = list(prime_ht_by_month.keys())
    prime_ht_data = list(prime_ht_by_month.values())
    prime_nette_data = list(prime_nette_by_month.values())

    return render(request, 'insurances/dashboard_prime_types.html', {
        'months': months,
        'prime_ht_data': prime_ht_data,
        'prime_nette_data': prime_nette_data,
    })


def sanlam_contracts(request):
    sanlam_company = Company.objects.get(name='SANLAM')

    if sanlam_company:
        sanlam_contracts = Contract.objects.filter(
            Q(company=sanlam_company)
        ).order_by('-id').distinct()
        sanlam_contracts_count = sanlam_contracts.count()
        paginator = Paginator(sanlam_contracts, 5)
        page = request.GET.get('page')
        sanlam_contracts = paginator.get_page(page)

    else:
        sanlam_contracts_count = 0
        sanlam_contracts = []

    return render(request, 'insurances/sanlam_contracts.html',
                  {'sanlam_contracts_count': sanlam_contracts_count,
                   'sanlam_contracts': sanlam_contracts,
                   'navbar': 'sanlam'})


def nsia_contracts(request):
    nsia_company = Company.objects.get(name='NSIA')

    if nsia_company:
        nsia_contracts = Contract.objects.filter(
            Q(company=nsia_company)
        ).order_by('-id').distinct()
        nsia_contracts_count = nsia_contracts.count()

        paginator = Paginator(nsia_contracts, 5)
        page = request.GET.get('page')
        nsia_contracts = paginator.get_page(page)

    else:
        nsia_contracts_count = 0
        nsia_contracts = []

    return render(request, 'insurances/nsia_contracts.html',
                  {'nsia_contracts_count': nsia_contracts_count,
                   'nsia_contracts': nsia_contracts,
                   'navbar': 'nsia'})


def fidelia_contracts(request):
    fidelia_company = Company.objects.get(name='FIDELIA')
    prime_fidelia = 0
    if fidelia_company:
        fidelia_contracts = Contract.objects.filter(
            Q(company=fidelia_company)
        ).order_by('-id').distinct()

        fidelia_contracts_count = fidelia_contracts.count()
        paginator = Paginator(fidelia_contracts, 5)
        page = request.GET.get('page')
        fidelia_contracts = paginator.get_page(page)

    else:
        fidelia_contracts_count = 0
        fidelia_contracts = []
        prime_fidelia = 0

    return render(request, 'insurances/fidelia_contracts.html',
                  {'fidelia_contracts_count': fidelia_contracts_count,
                   'fidelia_contracts': fidelia_contracts,
                   'navbar': 'fidelia',
                   'prime_fidelia': prime_fidelia})


def prime_company_dashboard(request):
    # Obtenez la liste des compagnies avec le nombre total de contrats et la somme totale des primes ttc
    companies_data = Company.objects.annotate(
        total_contracts=Count('contract'),
        total_prime_ttc=Sum('contract__prime_ttc')
    ).values('name', 'total_contracts', 'total_prime_ttc')

    # Calculez le pourcentage des contrats et des primes ttc par rapport au total
    total_contracts_sum = sum(company['total_contracts'] for company in companies_data)
    total_prime_ttc_sum = sum(company['total_prime_ttc'] for company in companies_data)

    for company in companies_data:
        company['total_contracts_percentage'] = (company['total_contracts'] / total_contracts_sum) * 100
        company['total_prime_ttc_percentage'] = (company['total_prime_ttc'] / total_prime_ttc_sum) * 100

    return render(request, 'insurances/dashboard.html', {
        'companies_data': companies_data,
    })


def prime_type_dashboard(request):
    # Obtenez la liste des compagnies avec le nombre total de contrats et la somme totale des primes ttc
    types_data = Type.objects.annotate(
        total_contracts=Count('contract'),
        total_prime_ttc=Sum('contract__prime_ttc')
    ).values('libelle', 'total_contracts', 'total_prime_ttc')

    # Calculez le pourcentage des contrats et des primes ttc par rapport au total
    total_contracts_sum = sum(type_contract['total_contracts'] for type_contract in types_data)
    total_prime_ttc_sum = sum(type_contract['total_prime_ttc'] for type_contract in types_data)

    for type_contract in types_data:
        type_contract['total_contracts_percentage'] = (type_contract['total_contracts'] / total_contracts_sum) * 100
        type_contract['total_prime_ttc_percentage'] = (type_contract['total_prime_ttc'] / total_prime_ttc_sum) * 100

    return render(request, 'insurances/dashboard_prime_types.html', {
        'types_data': types_data,
        'total_contracts': total_contracts_sum,
        'total_prime_ttc': total_prime_ttc_sum
    })


def prime_vehicle_dashboard_monthly(request):
    # Récupérer les contrats pour l'année en cours
    year = datetime.date.today().year
    contracts_year = Contract.objects.filter(register_date__year=year)
    contracts = contracts_year.filter(vehicle__isnull=False)
    # Initialiser un dictionnaire pour stocker les totaux mensuels
    monthly_totals = {}

    # Calculer les totaux mensuels
    for month in range(1, 13):
        # Filtrer les contrats pour le mois actuel
        month_contracts = contracts.filter(register_date__month=month)

        # Calculer le total des primes TTC pour ce mois
        total_prime_ttc = month_contracts.aggregate(Sum('prime_ttc'))['prime_ttc__sum'] or Decimal('0.00')

        # Calculer le total des primes nettes pour ce mois
        total_prime_net = month_contracts.aggregate(Sum('prime_net'))['prime_net__sum'] or Decimal('0.00')

        # Calculer le total des primes HT pour ce mois
        total_prime_ht = month_contracts.aggregate(Sum('prime_ht'))['prime_ht__sum'] or Decimal('0.00')

        # Calculer le total des accessoires pour ce mois
        total_accessoires = month_contracts.aggregate(Sum('accessory'))['accessory__sum'] or Decimal('0.00')

        # Calculer le total des taux pour ce mois
        total_taux = month_contracts.aggregate(Sum('rate'))['rate__sum'] or Decimal('0.00')

        # Calculer le total des commissions pour ce mois
        total_commission = month_contracts.aggregate(Sum('committee'))['committee__sum'] or Decimal('0.00')

        # Stocker les totaux dans le dictionnaire avec le mois comme clé
        monthly_totals[month] = {
            'total_prime_ttc': total_prime_ttc,
            'total_prime_net': total_prime_net,
            'total_prime_ht': total_prime_ht,
            'total_accessoires': total_accessoires,
            'total_taux': total_taux,
            'total_commission': total_commission,
        }
    return render(request, 'insurances/dashboard_vehicle.html', {'monthly_totals': monthly_totals})


def prime_person_dashboard_monthly(request):
    # Récupérer les contrats pour l'année en cours
    year = datetime.date.today().year
    contracts_year = Contract.objects.filter(register_date__year=year)
    contracts = contracts_year.filter(person__isnull=False)
    # Initialiser un dictionnaire pour stocker les totaux mensuels
    monthly_totals = {}

    # Calculer les totaux mensuels
    for month in range(1, 13):
        # Filtrer les contrats pour le mois actuel
        month_contracts = contracts.filter(register_date__month=month)

        # Calculer le total des primes TTC pour ce mois
        total_prime_ttc = month_contracts.aggregate(Sum('prime_ttc'))['prime_ttc__sum'] or Decimal('0.00')

        # Calculer le total des primes nettes pour ce mois
        total_prime_net = month_contracts.aggregate(Sum('prime_net'))['prime_net__sum'] or Decimal('0.00')

        # Calculer le total des primes HT pour ce mois
        total_prime_ht = month_contracts.aggregate(Sum('prime_ht'))['prime_ht__sum'] or Decimal('0.00')

        # Calculer le total des accessoires pour ce mois
        total_accessoires = month_contracts.aggregate(Sum('accessory'))['accessory__sum'] or Decimal('0.00')

        # Calculer le total des taux pour ce mois
        total_taux = month_contracts.aggregate(Sum('rate'))['rate__sum'] or Decimal('0.00')

        # Calculer le total des commissions pour ce mois
        total_commission = month_contracts.aggregate(Sum('committee'))['committee__sum'] or Decimal('0.00')

        # Stocker les totaux dans le dictionnaire avec le mois comme clé
        monthly_totals[month] = {
            'total_prime_ttc': total_prime_ttc,
            'total_prime_net': total_prime_net,
            'total_prime_ht': total_prime_ht,
            'total_accessoires': total_accessoires,
            'total_taux': total_taux,
            'total_commission': total_commission,
        }
    return render(request, 'insurances/dashboard_person.html', {'monthly_totals': monthly_totals})


def prime_dommage_dashboard_monthly(request):
    # Récupérer les contrats pour l'année en cours
    year = datetime.date.today().year
    contracts_year = Contract.objects.filter(register_date__year=year)
    contracts = contracts_year.filter(other__isnull=False)
    # Initialiser un dictionnaire pour stocker les totaux mensuels
    monthly_totals = {}

    # Calculer les totaux mensuels
    for month in range(1, 13):
        # Filtrer les contrats pour le mois actuel
        month_contracts = contracts.filter(register_date__month=month)

        # Calculer le total des primes TTC pour ce mois
        total_prime_ttc = month_contracts.aggregate(Sum('prime_ttc'))['prime_ttc__sum'] or Decimal('0.00')

        # Calculer le total des primes nettes pour ce mois
        total_prime_net = month_contracts.aggregate(Sum('prime_net'))['prime_net__sum'] or Decimal('0.00')

        # Calculer le total des primes HT pour ce mois
        total_prime_ht = month_contracts.aggregate(Sum('prime_ht'))['prime_ht__sum'] or Decimal('0.00')

        # Calculer le total des accessoires pour ce mois
        total_accessoires = month_contracts.aggregate(Sum('accessory'))['accessory__sum'] or Decimal('0.00')

        # Calculer le total des taux pour ce mois
        total_taux = month_contracts.aggregate(Sum('rate'))['rate__sum'] or Decimal('0.00')

        # Calculer le total des commissions pour ce mois
        total_commission = month_contracts.aggregate(Sum('committee'))['committee__sum'] or Decimal('0.00')

        # Stocker les totaux dans le dictionnaire avec le mois comme clé
        monthly_totals[month] = {
            'total_prime_ttc': total_prime_ttc,
            'total_prime_net': total_prime_net,
            'total_prime_ht': total_prime_ht,
            'total_accessoires': total_accessoires,
            'total_taux': total_taux,
            'total_commission': total_commission,
        }
    return render(request, 'insurances/dashboard_dommage.html', {'monthly_totals': monthly_totals})


def prime_dashboard(request):
    # Appelez les trois vues existantes pour obtenir leurs données
    vehicle_data = prime_vehicle_dashboard_monthly(request)
    person_data = prime_person_dashboard_monthly(request)
    dommage_data = prime_dommage_dashboard_monthly(request)

    # Passez les données au template
    return render(request, 'insurances/my_dash.html', {
        'vehicle_data': vehicle_data,
        'person_data': person_data,
        'dommage_data': dommage_data,
    })


def prime_company_dashboard_monthly(request):
    # Récupérer les contrats pour l'année en cours
    year = datetime.date.today().year
    contracts = Contract.objects.filter(register_date__year=year)

    # Initialiser un dictionnaire pour stocker les totaux mensuels
    monthly_totals = {}

    # Calculer les totaux mensuels
    for month in range(1, 13):
        # Filtrer les contrats pour le mois actuel
        month_contracts = contracts.filter(register_date__month=month)

        # Calculer le total des primes TTC pour ce mois
        total_prime_ttc = month_contracts.aggregate(Sum('prime_ttc'))['prime_ttc__sum'] or Decimal('0.00')

        # Calculer le total des primes nettes pour ce mois
        total_prime_net = month_contracts.aggregate(Sum('prime_net'))['prime_net__sum'] or Decimal('0.00')

        # Calculer le total des primes HT pour ce mois
        total_prime_ht = month_contracts.aggregate(Sum('prime_ht'))['prime_ht__sum'] or Decimal('0.00')

        # Calculer le total des accessoires pour ce mois
        total_accessoires = month_contracts.aggregate(Sum('accessory'))['accessory__sum'] or Decimal('0.00')

        # Calculer le total des taux pour ce mois
        total_taux = month_contracts.aggregate(Sum('rate'))['rate__sum'] or Decimal('0.00')

        # Calculer le total des commissions pour ce mois
        total_commission = month_contracts.aggregate(Sum('committee'))['committee__sum'] or Decimal('0.00')

        # Stocker les totaux dans le dictionnaire avec le mois comme clé
        monthly_totals[month] = {
            'total_prime_ttc': total_prime_ttc,
            'total_prime_net': total_prime_net,
            'total_prime_ht': total_prime_ht,
            'total_accessoires': total_accessoires,
            'total_taux': total_taux,
            'total_commission': total_commission,
        }

    # Passer les données au template
    return render(request, 'insurances/dashboard_base.html', {'monthly_totals': monthly_totals})


# --------------------------------------------------------------------------------------#
def contracts_dashboard2(request):
    # Extraction des données pertinentes pour les graphiques
    policies = Policy.objects.all()
    policy_labels = [policy.guaranttee for policy in policies]
    policy_counts = [Contract.objects.filter(policy=policy).count() for policy in policies]

    contracts = Contract.objects.all()
    contract_dates = [contract.effect for contract in contracts]
    prime_ht_values = [contract.prime_ht for contract in contracts]

    # Création du graphique de répartition des garanties
    plt.figure(figsize=(10, 6))
    plt.bar(policy_labels, policy_counts)
    plt.xlabel('Garanties')
    plt.ylabel('Nombre de contrats')
    plt.title('Répartition des garanties dans les contrats')
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Enregistrement du graphique dans un fichier
    # guarantee_chart_path = 'static/guarantee_chart.png'
    # plt.savefig(guarantee_chart_path)
    plt.savefig('static/images/guarantee_chart.png')
    plt.close()

    # Création du graphique d'évolution des primes au fil du temps
    plt.figure(figsize=(10, 6))
    plt.plot(contract_dates, prime_ht_values, marker='o')
    plt.xlabel('Date de contrat')
    plt.ylabel('Prime HT')
    plt.title('Évolution des primes au fil du temps')
    plt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Enregistrement du graphique dans un fichier
    # prime_chart_path = 'static/prime_chart.png'
    # plt.savefig(prime_chart_path)
    plt.savefig('static/images/prime_chart.png')
    plt.close()

    return render(request, 'insurances/dashboard.html', {
        'guarantee_chart_path': 'static/images/guarantee_chart.png',
        'prime_chart_path': 'static/images/prime_chart.png',
    })

# def contracts_dashboard_policy(request):
#     # Extraction des données pertinentes pour les graphiques
#     policy_counts = Contract.objects.values('guaranttee__libelle').annotate(count=Count('guaranttee')).order_by(
#         '-count')
#     policy_labels = [item['guaranttee__libelle'] for item in policy_counts]
#     policy_counts = [item['count'] for item in policy_counts]
#
#     return render(request, 'insurances/dashboard.html', {
#         'policy_labels': policy_labels,
#         'policy_counts': policy_counts,
#     })
